from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


PROJECT_ROOT = Path(r"C:\01. GitHub\BeSmart15dModel")
WORKING_COPY_DIR = PROJECT_ROOT / "01_RAW" / "WORKING_COPY"
CLONE_DIR = PROJECT_ROOT / "08_CLONE"
EXPORTS_DIR = PROJECT_ROOT / "09_EXPORTS"
LOGS_DIR = PROJECT_ROOT / "06_TOOLS" / "Logs"

SUMMARY_JSON = EXPORTS_DIR / "CLONE.STAGE6.FLUJO_MENSUAL.SUMMARY.json"
RUN_LOG = LOGS_DIR / "rebuild_flujo_mensual_ia_stage6_normalized.log"

TARGET_SHEET = "Flujo mensual.IA"

MAP = {
    "rent_market_plus_gc": "Q4",
    "furniture_capex": "Q7",
    "contribuciones": "Q10",
    "dividendo_total_credito": "Q13",
    "gasto_operacional": "Q16",
    "corretaje_arriendo": "Q19",
    "imprevistos": "Q28",
    "administracion": "Q31",
}

OCCUPANCY_RATE = 0.90

START_ROW = 2
END_ROW = 28
START_COL = 1
END_COL = 6


def now_utc() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def log(msg: str) -> None:
    with RUN_LOG.open("a", encoding="utf-8") as f:
        f.write(f"[{now_utc()}] {msg}\n")


def reset_log() -> None:
    RUN_LOG.write_text("", encoding="utf-8")


def parse_float(v) -> float:
    if v is None:
        return 0.0
    txt = str(v).strip()
    if txt == "":
        return 0.0
    try:
        txt = txt.replace(",", "")
        return float(txt)
    except Exception:
        return 0.0


def get_latest_stage5_clone() -> Path:
    files = sorted(CLONE_DIR.glob("*.stage5.IA.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"No se encontró ningún archivo *.stage5.IA.xlsx en {CLONE_DIR}")
    return files[0]


def get_latest_working_copy() -> Path:
    files = sorted(WORKING_COPY_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"No se encontró ningún working copy .xlsx en {WORKING_COPY_DIR}")
    return files[0]


def get_calc_value(ws, cell_address: str) -> dict:
    raw = ws[cell_address].value
    return {
        "cell_address": cell_address,
        "value_raw": "" if raw is None else str(raw),
        "value_num": parse_float(raw)
    }


def clear_block(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).value = None


def write_rows(ws, start_row: int, rows: list[list]) -> None:
    for i, row in enumerate(rows):
        for j, value in enumerate(row, start=1):
            ws.cell(row=start_row + i, column=j).value = value


def style_block(ws, start_row: int, row_count: int) -> None:
    blue = PatternFill(fill_type="solid", fgColor="D9EAF7")
    yellow = PatternFill(fill_type="solid", fgColor="FFF2CC")
    green = PatternFill(fill_type="solid", fgColor="E2F0D9")
    orange = PatternFill(fill_type="solid", fgColor="FCE4D6")

    for c in range(1, 7):
        ws.cell(row=start_row, column=c).fill = blue
        ws.cell(row=start_row + 1, column=c).fill = blue

    for c in range(1, 7):
        ws.cell(row=start_row + 6, column=c).fill = green

    for c in range(1, 7):
        ws.cell(row=start_row + 9, column=c).fill = orange

    ws.cell(row=start_row + row_count - 1, column=1).fill = yellow
    ws.cell(row=start_row + row_count - 1, column=2).fill = yellow


def apply_number_formats(ws, start_row: int) -> None:
    currency_rows = [start_row + 2, start_row + 4, start_row + 5, start_row + 6, start_row + 10]
    ratio_rows = [start_row + 3]

    for r in currency_rows:
        ws.cell(row=r, column=2).number_format = '#,##0.00'

    for r in ratio_rows:
        ws.cell(row=r, column=2).number_format = '0.00%'


def main() -> int:
    reset_log()
    log("Inicio rebuild_flujo_mensual_ia_stage6_normalized.")

    try:
        source_stage5 = get_latest_stage5_clone()
        working_copy = get_latest_working_copy()

        log(f"Stage5 clone seleccionado: {source_stage5}")
        log(f"Working copy fuente data_only: {working_copy}")

        wb_calc = load_workbook(working_copy, data_only=True)
        if "Arriendo" not in wb_calc.sheetnames:
            raise ValueError("No existe hoja 'Arriendo' en working copy")

        ws_arr = wb_calc["Arriendo"]

        rent = get_calc_value(ws_arr, MAP["rent_market_plus_gc"])
        furniture = get_calc_value(ws_arr, MAP["furniture_capex"])
        contrib = get_calc_value(ws_arr, MAP["contribuciones"])
        dividend_total = get_calc_value(ws_arr, MAP["dividendo_total_credito"])
        opex = get_calc_value(ws_arr, MAP["gasto_operacional"])
        broker = get_calc_value(ws_arr, MAP["corretaje_arriendo"])
        unexpected = get_calc_value(ws_arr, MAP["imprevistos"])
        admin = get_calc_value(ws_arr, MAP["administracion"])

        # Stage 6: solo bloque operativo mensual comparable
        monthly_rent_base = rent["value_num"]
        monthly_operating_expense = (
            contrib["value_num"]
            + opex["value_num"]
            + broker["value_num"]
            + unexpected["value_num"]
            + admin["value_num"]
        )
        monthly_income_effective = monthly_rent_base * OCCUPANCY_RATE
        monthly_operating_net = monthly_income_effective - monthly_operating_expense

        log(f"Q4 rent raw={rent['value_raw']} num={monthly_rent_base}")
        log(f"Q13 financing raw={dividend_total['value_raw']} num={dividend_total['value_num']} -> excluded from operating monthly net")
        log(f"Operating expense total={monthly_operating_expense}")

        wb_out = load_workbook(source_stage5)
        if TARGET_SHEET not in wb_out.sheetnames:
            raise ValueError(f"No existe hoja objetivo: {TARGET_SHEET}")

        ws = wb_out[TARGET_SHEET]

        rows = [
            ["STAGE6 - MINIMAL NORMALIZATION", "", "", "", "", ""],
            ["Concepto", "Valor", "Unidad", "Origen", "Comentario", "Estado"],
            ["Arriendo base + GG", monthly_rent_base, "CLP?", "Arriendo!Q4 data_only", "Ingreso base calculado", "OK"],
            ["Ocupación estabilizada", OCCUPANCY_RATE, "ratio", "PARAM", "Supuesto confirmado", "OK"],
            ["Ingreso mensual efectivo", monthly_income_effective, "CLP?", "ENGINE_STAGE6", "Q4 x 90%", "OK"],
            ["Egreso operativo mensual comparable", monthly_operating_expense, "MIX_PENDING", "Q10+Q16+Q19+Q28+Q31", "Sin incluir Q13", "PENDING_UNIT_VALIDATION"],
            ["Flujo operativo mensual neto", monthly_operating_net, "MIX_PENDING", "ENGINE_STAGE6", "Ingreso efectivo - egreso operativo comparable", "OK"],
            ["CAPEX amoblado excluido", furniture["value_num"], "CLP", "Arriendo!Q7", "No entra a flujo mensual", "INFO"],
            ["FINANCIAMIENTO / NORMALIZACION PENDIENTE", "", "", "", "", ""],
            ["Dividendo total crédito", dividend_total["value_num"], "UF/CLP_MIX_PENDING", "Arriendo!Q13", "No se suma al neto mensual hasta normalizar unidades", "PENDING_NORMALIZATION"],
            ["Nota 1", "Workbook mezcla UF y CLP.", "", "USER_CONTEXT", "", "INFO"],
            ["Nota 2", "Flujo 5 años C12 opera por años.", "", "USER_CONTEXT", "", "INFO"],
            ["Nota 3", "Granularidad por dpto vs total sigue pendiente de cierre completo.", "", "SYSTEM", "", "INFO"],
        ]

        clear_block(ws, START_ROW, END_ROW, START_COL, END_COL)
        write_rows(ws, START_ROW, rows)
        style_block(ws, START_ROW, len(rows))
        apply_number_formats(ws, START_ROW)

        out_name = source_stage5.name.replace(".stage5.IA.xlsx", ".stage6.IA.xlsx")
        out_path = CLONE_DIR / out_name
        wb_out.save(out_path)

        summary = {
            "timestamp_utc": now_utc(),
            "source_stage5_clone": str(source_stage5),
            "source_working_copy_data_only": str(working_copy),
            "output_stage6_clone": str(out_path),
            "target_sheet": TARGET_SHEET,
            "normalization_rule": {
                "included_in_operating_monthly_flow": [
                    "Arriendo!Q4",
                    "Arriendo!Q10",
                    "Arriendo!Q16",
                    "Arriendo!Q19",
                    "Arriendo!Q28",
                    "Arriendo!Q31"
                ],
                "excluded_pending_normalization": [
                    "Arriendo!Q13"
                ],
                "excluded_capex": [
                    "Arriendo!Q7"
                ]
            },
            "raw_values": {
                "Q4": rent["value_raw"],
                "Q7": furniture["value_raw"],
                "Q10": contrib["value_raw"],
                "Q13": dividend_total["value_raw"],
                "Q16": opex["value_raw"],
                "Q19": broker["value_raw"],
                "Q28": unexpected["value_raw"],
                "Q31": admin["value_raw"]
            },
            "values": {
                "monthly_rent_base": monthly_rent_base,
                "monthly_operating_expense": monthly_operating_expense,
                "monthly_income_effective": monthly_income_effective,
                "monthly_operating_net": monthly_operating_net,
                "excluded_furniture_capex": furniture["value_num"],
                "excluded_financing_pending_normalization": dividend_total["value_num"]
            },
            "status": "OK_STAGE6_MINIMAL_NORMALIZATION"
        }

        with SUMMARY_JSON.open("w", encoding="utf-8") as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)

        log(f"Workbook stage6 guardado: {out_path}")
        log("Proceso completado OK.")
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0

    except Exception as e:
        log(f"ERROR: {type(e).__name__}: {e}")
        print(f"ERROR: {type(e).__name__}: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
