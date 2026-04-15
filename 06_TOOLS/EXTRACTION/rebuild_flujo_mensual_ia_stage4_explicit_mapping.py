from __future__ import annotations

import json
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


PROJECT_ROOT = Path(r"C:\01. GitHub\BeSmart15dModel")
DATA_DIR = PROJECT_ROOT / "01_DATA"
CLONE_DIR = PROJECT_ROOT / "08_CLONE"
EXPORTS_DIR = PROJECT_ROOT / "09_EXPORTS"
LOGS_DIR = PROJECT_ROOT / "06_TOOLS" / "Logs"

DB_PATH = DATA_DIR / "BeSmart15dModel.stage1.sqlite"
SUMMARY_JSON = EXPORTS_DIR / "CLONE.STAGE4.FLUJO_MENSUAL.SUMMARY.json"
RUN_LOG = LOGS_DIR / "rebuild_flujo_mensual_ia_stage4_explicit_mapping.log"

TARGET_SHEET = "Flujo mensual.IA"

# ============================================================
# EXPLICIT MAPPING - ARRIENDO
# ============================================================
MAP = {
    "rent_market_plus_gc": "Q4",
    "furniture_capex": "Q7",
    "contribuciones": "Q10",
    "dividendo": "Q13",
    "gasto_operacional": "Q16",
    "corretaje_arriendo": "Q19",
    "imprevistos": "Q28",
    "administracion": "Q31",
}

OCCUPANCY_RATE = 0.90

START_ROW = 2
END_ROW = 20
START_COL = 1
END_COL = 6


def now_utc() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def log(msg: str) -> None:
    with RUN_LOG.open("a", encoding="utf-8") as f:
        f.write(f"[{now_utc()}] {msg}\n")


def reset_log() -> None:
    RUN_LOG.write_text("", encoding="utf-8")


def parse_float(v) -> float:
    if v is None:
        return 0.0
    txt = str(v).strip()
    if txt == "":
        return 0.0
    try:
        txt = txt.replace(",", "")
        return float(txt)
    except Exception:
        return 0.0


def get_latest_stage3_clone() -> Path:
    files = sorted(CLONE_DIR.glob("*.stage3.IA.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"No se encontró ningún archivo *.stage3.IA.xlsx en {CLONE_DIR}")
    return files[0]


def get_render_value(conn: sqlite3.Connection, sheet_name: str, cell_address: str) -> dict:
    df = pd.read_sql_query(f"""
        SELECT sheet_name, cell_address, value, formula, number_format, data_type
        FROM render_cells
        WHERE sheet_name = '{sheet_name}'
          AND cell_address = '{cell_address}'
        LIMIT 1
    """, conn)

    if df.empty:
        return {
            "sheet_name": sheet_name,
            "cell_address": cell_address,
            "value_raw": "",
            "value_num": 0.0,
            "found": False
        }

    raw = df.iloc[0]["value"]
    return {
        "sheet_name": str(df.iloc[0]["sheet_name"]),
        "cell_address": str(df.iloc[0]["cell_address"]),
        "value_raw": "" if pd.isna(raw) else str(raw),
        "value_num": parse_float(raw),
        "found": True
    }


def clear_block(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).value = None


def write_rows(ws, start_row: int, rows: list[list]) -> None:
    for i, row in enumerate(rows):
        for j, value in enumerate(row, start=1):
            ws.cell(row=start_row + i, column=j).value = value


def style_block(ws, start_row: int, row_count: int) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    info_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    ok_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")

    for c in range(1, 7):
        ws.cell(row=start_row, column=c).fill = header_fill
        ws.cell(row=start_row + 1, column=c).fill = header_fill

    for c in range(1, 7):
        ws.cell(row=start_row + 6, column=c).fill = ok_fill

    ws.cell(row=start_row + row_count - 1, column=1).fill = info_fill
    ws.cell(row=start_row + row_count - 1, column=2).fill = info_fill


def apply_number_formats(ws, start_row: int) -> None:
    currency_rows = [start_row + 2, start_row + 4, start_row + 5, start_row + 6]
    ratio_rows = [start_row + 3]

    for r in currency_rows:
        ws.cell(row=r, column=2).number_format = '#,##0.00'

    for r in ratio_rows:
        ws.cell(row=r, column=2).number_format = '0.00%'


def main() -> int:
    reset_log()
    log("Inicio rebuild_flujo_mensual_ia_stage4_explicit_mapping.")

    try:
        if not DB_PATH.exists():
            raise FileNotFoundError(f"No existe DB: {DB_PATH}")

        stage3_path = get_latest_stage3_clone()
        log(f"Stage3 clone seleccionado: {stage3_path}")

        conn = sqlite3.connect(DB_PATH)
        try:
            rent = get_render_value(conn, "Arriendo", MAP["rent_market_plus_gc"])
            furniture = get_render_value(conn, "Arriendo", MAP["furniture_capex"])
            contrib = get_render_value(conn, "Arriendo", MAP["contribuciones"])
            dividend = get_render_value(conn, "Arriendo", MAP["dividendo"])
            opex = get_render_value(conn, "Arriendo", MAP["gasto_operacional"])
            broker = get_render_value(conn, "Arriendo", MAP["corretaje_arriendo"])
            unexpected = get_render_value(conn, "Arriendo", MAP["imprevistos"])
            admin = get_render_value(conn, "Arriendo", MAP["administracion"])
        finally:
            conn.close()

        monthly_rent_base = rent["value_num"]
        monthly_expense = (
            contrib["value_num"]
            + dividend["value_num"]
            + opex["value_num"]
            + broker["value_num"]
            + unexpected["value_num"]
            + admin["value_num"]
        )
        monthly_income_effective = monthly_rent_base * OCCUPANCY_RATE
        monthly_net = monthly_income_effective - monthly_expense
        annual_income_effective = monthly_income_effective * 12

        log(f"Rent base: {monthly_rent_base} desde Arriendo!{rent['cell_address']}")
        log(f"Expense total: {monthly_expense}")
        log(f"Furniture CAPEX excluded from monthly flow: {furniture['value_num']} desde Arriendo!{furniture['cell_address']}")

        wb = load_workbook(stage3_path)

        if TARGET_SHEET not in wb.sheetnames:
            raise ValueError(f"No existe hoja objetivo: {TARGET_SHEET}")

        ws = wb[TARGET_SHEET]

        rows = [
            ["STAGE4 - EXPLICIT BUSINESS MAPPING", "", "", "", "", ""],
            ["Concepto", "Valor", "Unidad", "Origen", "Comentario", "Estado"],
            ["Arriendo base + GG", monthly_rent_base, "CLP", "Arriendo!Q4", "Ingreso mensual base correcto", "OK"],
            ["Ocupación estabilizada", OCCUPANCY_RATE, "ratio", "PARAM", "Supuesto confirmado", "OK"],
            ["Ingreso mensual efectivo", monthly_income_effective, "CLP", "ENGINE_STAGE4", "Q4 x 90%", "OK"],
            ["Egreso mensual total", monthly_expense, "CLP", "Q10+Q13+Q16+Q19+Q28+Q31", "Suma explícita de egresos", "OK"],
            ["Flujo mensual neto", monthly_net, "CLP", "ENGINE_STAGE4", "Ingreso efectivo - egreso", "OK"],
            ["Ingreso anual estabilizado", annual_income_effective, "CLP", "ENGINE_STAGE4", "Mensual efectivo x 12", "OK"],
            ["CAPEX amoblado excluido", furniture["value_num"], "CLP", "Arriendo!Q7", "No entra al flujo mensual", "INFO"],
            ["Nota", "Stage 4 elimina heurística y usa mapping explícito de negocio.", "", "SYSTEM", "", "INFO"],
        ]

        clear_block(ws, START_ROW, END_ROW, START_COL, END_COL)
        write_rows(ws, START_ROW, rows)
        style_block(ws, START_ROW, len(rows))
        apply_number_formats(ws, START_ROW)

        out_name = stage3_path.name.replace(".stage3.IA.xlsx", ".stage4.IA.xlsx")
        out_path = CLONE_DIR / out_name
        wb.save(out_path)

        summary = {
            "timestamp_utc": now_utc(),
            "source_stage3_clone": str(stage3_path),
            "output_stage4_clone": str(out_path),
            "target_sheet": TARGET_SHEET,
            "explicit_mapping": {
                "rent_market_plus_gc": "Arriendo!Q4",
                "furniture_capex_excluded": "Arriendo!Q7",
                "contribuciones": "Arriendo!Q10",
                "dividendo": "Arriendo!Q13",
                "gasto_operacional": "Arriendo!Q16",
                "corretaje_arriendo": "Arriendo!Q19",
                "imprevistos": "Arriendo!Q28",
                "administracion": "Arriendo!Q31"
            },
            "values": {
                "monthly_rent_base": monthly_rent_base,
                "monthly_expense": monthly_expense,
                "monthly_income_effective": monthly_income_effective,
                "monthly_net": monthly_net,
                "annual_income_effective": annual_income_effective,
                "excluded_furniture_capex": furniture["value_num"]
            },
            "status": "OK_STAGE4_EXPLICIT_BUSINESS_MAPPING"
        }

        with SUMMARY_JSON.open("w", encoding="utf-8") as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)

        log(f"Workbook stage4 guardado: {out_path}")
        log("Proceso completado OK.")
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0

    except Exception as e:
        log(f"ERROR: {type(e).__name__}: {e}")
        print(f"ERROR: {type(e).__name__}: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
