from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


PROJECT_ROOT = Path(r"C:\01. GitHub\BeSmart15dModel")
WORKING_COPY_DIR = PROJECT_ROOT / "01_RAW" / "WORKING_COPY"
CLONE_DIR = PROJECT_ROOT / "08_CLONE"
EXPORTS_DIR = PROJECT_ROOT / "09_EXPORTS"
LOGS_DIR = PROJECT_ROOT / "06_TOOLS" / "Logs"

SUMMARY_JSON = EXPORTS_DIR / "CLONE.STAGE5.FLUJO_MENSUAL.SUMMARY.json"
RUN_LOG = LOGS_DIR / "rebuild_flujo_mensual_ia_stage5_data_only.log"

TARGET_SHEET = "Flujo mensual.IA"

# ============================================================
# EXPLICIT BUSINESS MAPPING - ARRIENDO
# ============================================================
MAP = {
    "rent_market_plus_gc": "Q4",
    "furniture_capex": "Q7",
    "contribuciones": "Q10",
    "dividendo": "Q13",
    "gasto_operacional": "Q16",
    "corretaje_arriendo": "Q19",
    "imprevistos": "Q28",
    "administracion": "Q31",
}

OCCUPANCY_RATE = 0.90

START_ROW = 2
END_ROW = 20
START_COL = 1
END_COL = 6


def now_utc() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def log(msg: str) -> None:
    with RUN_LOG.open("a", encoding="utf-8") as f:
        f.write(f"[{now_utc()}] {msg}\n")


def reset_log() -> None:
    RUN_LOG.write_text("", encoding="utf-8")


def parse_float(v) -> float:
    if v is None:
        return 0.0
    txt = str(v).strip()
    if txt == "":
        return 0.0
    try:
        txt = txt.replace(",", "")
        return float(txt)
    except Exception:
        return 0.0


def get_latest_stage4_clone() -> Path:
    files = sorted(CLONE_DIR.glob("*.stage4.IA.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"No se encontró ningún archivo *.stage4.IA.xlsx en {CLONE_DIR}")
    return files[0]


def get_latest_working_copy() -> Path:
    files = sorted(WORKING_COPY_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"No se encontró ningún working copy .xlsx en {WORKING_COPY_DIR}")
    return files[0]


def get_calc_value(ws, cell_address: str) -> dict:
    raw = ws[cell_address].value
    return {
        "cell_address": cell_address,
        "value_raw": "" if raw is None else str(raw),
        "value_num": parse_float(raw)
    }


def clear_block(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).value = None


def write_rows(ws, start_row: int, rows: list[list]) -> None:
    for i, row in enumerate(rows):
        for j, value in enumerate(row, start=1):
            ws.cell(row=start_row + i, column=j).value = value


def style_block(ws, start_row: int, row_count: int) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    info_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    ok_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")

    for c in range(1, 7):
        ws.cell(row=start_row, column=c).fill = header_fill
        ws.cell(row=start_row + 1, column=c).fill = header_fill

    for c in range(1, 7):
        ws.cell(row=start_row + 6, column=c).fill = ok_fill

    ws.cell(row=start_row + row_count - 1, column=1).fill = info_fill
    ws.cell(row=start_row + row_count - 1, column=2).fill = info_fill


def apply_number_formats(ws, start_row: int) -> None:
    currency_rows = [start_row + 2, start_row + 4, start_row + 5, start_row + 6, start_row + 7]
    ratio_rows = [start_row + 3]

    for r in currency_rows:
        ws.cell(row=r, column=2).number_format = '#,##0.00'

    for r in ratio_rows:
        ws.cell(row=r, column=2).number_format = '0.00%'


def main() -> int:
    reset_log()
    log("Inicio rebuild_flujo_mensual_ia_stage5_data_only.")

    try:
        source_stage4 = get_latest_stage4_clone()
        working_copy = get_latest_working_copy()

        log(f"Stage4 clone seleccionado: {source_stage4}")
        log(f"Working copy fuente para data_only: {working_copy}")

        # Workbook fuente con valores calculados
        wb_calc = load_workbook(working_copy, data_only=True)
        if "Arriendo" not in wb_calc.sheetnames:
            raise ValueError("No existe hoja 'Arriendo' en working copy")

        ws_arr = wb_calc["Arriendo"]

        rent = get_calc_value(ws_arr, MAP["rent_market_plus_gc"])
        furniture = get_calc_value(ws_arr, MAP["furniture_capex"])
        contrib = get_calc_value(ws_arr, MAP["contribuciones"])
        dividend = get_calc_value(ws_arr, MAP["dividendo"])
        opex = get_calc_value(ws_arr, MAP["gasto_operacional"])
        broker = get_calc_value(ws_arr, MAP["corretaje_arriendo"])
        unexpected = get_calc_value(ws_arr, MAP["imprevistos"])
        admin = get_calc_value(ws_arr, MAP["administracion"])

        monthly_rent_base = rent["value_num"]
        monthly_expense = (
            contrib["value_num"]
            + dividend["value_num"]
            + opex["value_num"]
            + broker["value_num"]
            + unexpected["value_num"]
            + admin["value_num"]
        )
        monthly_income_effective = monthly_rent_base * OCCUPANCY_RATE
        monthly_net = monthly_income_effective - monthly_expense
        annual_income_effective = monthly_income_effective * 12

        log(f"Q4 (rent market + gc): raw={rent['value_raw']} num={monthly_rent_base}")
        log(f"Q7 (furniture capex): raw={furniture['value_raw']} num={furniture['value_num']}")
        log(f"Q10/Q13/Q16/Q19/Q28/Q31 expense total={monthly_expense}")

        wb_out = load_workbook(source_stage4)
        if TARGET_SHEET not in wb_out.sheetnames:
            raise ValueError(f"No existe hoja objetivo: {TARGET_SHEET}")

        ws = wb_out[TARGET_SHEET]

        rows = [
            ["STAGE5 - DATA_ONLY BUSINESS MAPPING", "", "", "", "", ""],
            ["Concepto", "Valor", "Unidad", "Origen", "Comentario", "Estado"],
            ["Arriendo base + GG", monthly_rent_base, "CLP", "Arriendo!Q4 data_only", "Valor calculado real", "OK"],
            ["Ocupación estabilizada", OCCUPANCY_RATE, "ratio", "PARAM", "Supuesto confirmado", "OK"],
            ["Ingreso mensual efectivo", monthly_income_effective, "CLP", "ENGINE_STAGE5", "Q4 x 90%", "OK"],
            ["Egreso mensual total", monthly_expense, "CLP", "Q10+Q13+Q16+Q19+Q28+Q31 data_only", "Suma explícita de egresos", "OK"],
            ["Flujo mensual neto", monthly_net, "CLP", "ENGINE_STAGE5", "Ingreso efectivo - egreso", "OK"],
            ["Ingreso anual estabilizado", annual_income_effective, "CLP", "ENGINE_STAGE5", "Mensual efectivo x 12", "OK"],
            ["CAPEX amoblado excluido", furniture["value_num"], "CLP", "Arriendo!Q7 data_only", "No entra al flujo mensual", "INFO"],
            ["Nota", "Stage 5 usa valores calculados del workbook (data_only=True).", "", "SYSTEM", "", "INFO"],
        ]

        clear_block(ws, START_ROW, END_ROW, START_COL, END_COL)
        write_rows(ws, START_ROW, rows)
        style_block(ws, START_ROW, len(rows))
        apply_number_formats(ws, START_ROW)

        out_name = source_stage4.name.replace(".stage4.IA.xlsx", ".stage5.IA.xlsx")
        out_path = CLONE_DIR / out_name
        wb_out.save(out_path)

        summary = {
            "timestamp_utc": now_utc(),
            "source_stage4_clone": str(source_stage4),
            "source_working_copy_data_only": str(working_copy),
            "output_stage5_clone": str(out_path),
            "target_sheet": TARGET_SHEET,
            "explicit_mapping": {
                "rent_market_plus_gc": "Arriendo!Q4",
                "furniture_capex_excluded": "Arriendo!Q7",
                "contribuciones": "Arriendo!Q10",
                "dividendo": "Arriendo!Q13",
                "gasto_operacional": "Arriendo!Q16",
                "corretaje_arriendo": "Arriendo!Q19",
                "imprevistos": "Arriendo!Q28",
                "administracion": "Arriendo!Q31"
            },
            "raw_values": {
                "Q4": rent["value_raw"],
                "Q7": furniture["value_raw"],
                "Q10": contrib["value_raw"],
                "Q13": dividend["value_raw"],
                "Q16": opex["value_raw"],
                "Q19": broker["value_raw"],
                "Q28": unexpected["value_raw"],
                "Q31": admin["value_raw"]
            },
            "values": {
                "monthly_rent_base": monthly_rent_base,
                "monthly_expense": monthly_expense,
                "monthly_income_effective": monthly_income_effective,
                "monthly_net": monthly_net,
                "annual_income_effective": annual_income_effective,
                "excluded_furniture_capex": furniture["value_num"]
            },
            "status": "OK_STAGE5_DATA_ONLY_BUSINESS_MAPPING"
        }

        with SUMMARY_JSON.open("w", encoding="utf-8") as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)

        log(f"Workbook stage5 guardado: {out_path}")
        log("Proceso completado OK.")
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0

    except Exception as e:
        log(f"ERROR: {type(e).__name__}: {e}")
        print(f"ERROR: {type(e).__name__}: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
