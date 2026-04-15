from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

PROJECT_ROOT = Path(r"C:\01. GitHub\BeSmart15dModel")
WORKING_COPY_DIR = PROJECT_ROOT / "01_RAW" / "WORKING_COPY"
LOGS_DIR = PROJECT_ROOT / "06_TOOLS" / "Logs"
EXPORTS_DIR = PROJECT_ROOT / "09_EXPORTS"

CORE_DIR = PROJECT_ROOT / "02_AUDIT" / "CORE"
CORE_INPUTS = CORE_DIR / "CORE.INPUTS.csv"
CORE_OUTPUTS = CORE_DIR / "CORE.OUTPUTS.csv"
CORE_FORMULAS = CORE_DIR / "CORE.FORMULAS.csv"
CORE_CONSTANTES = CORE_DIR / "CORE.CONSTANTES.csv"
CORE_DEPENDENCIAS = CORE_DIR / "CORE.DEPENDENCIAS.csv"
CORE_SUMMARY = EXPORTS_DIR / "CORE.EXTRACT.SUMMARY.json"

TARGET_SHEETS = [
    "Arriendo",
    "Flujo mensual",
    "Flujo 5 años C12",
    "Flujo 5 años C12 (2)",
    "Flujo 5 años C20",
]

KEYWORDS_OUTPUT = [
    "ingresos", "total ingresos", "egresos", "total egresos", "flujo",
    "flujo acumulado", "flujo caja anual", "flujo anual acumulado",
    "flujo financiamiento", "venta", "valor venta", "resultado"
]

KEYWORDS_INPUT = [
    "arriendo", "ocupación", "ocupacion", "vacancia", "precio", "descuento",
    "tasa", "credito", "crédito", "pie", "canon", "venta"
]

def now_utc():
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

def safe_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v)

def ensure_dirs():
    for p in [CORE_DIR, LOGS_DIR, EXPORTS_DIR]:
        p.mkdir(parents=True, exist_ok=True)

def get_latest_working_copy() -> Path:
    files = sorted(WORKING_COPY_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"No se encontraron .xlsx en {WORKING_COPY_DIR}")
    return files[0]

def normalize(s: Any) -> str:
    return safe_str(s).strip().lower()

def text_has_keyword(text: Any, keywords: list[str]) -> bool:
    t = normalize(text)
    return any(k in t for k in keywords)

def extract_refs(formula: str, current_sheet: str) -> list[tuple[str, str]]:
    refs = []
    if not formula or not formula.startswith("="):
        return refs

    pattern = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_\. ]+))!([$]?[A-Z]{1,3}[$]?\d+(?::[$]?[A-Z]{1,3}[$]?\d+)?)")
    for m in pattern.finditer(formula):
        sheet = (m.group(1) or m.group(2) or current_sheet).strip()
        addr = m.group(3)
        refs.append((sheet, addr))

    if "!" not in formula:
        pattern_local = re.compile(r"(?<![A-Za-z0-9_])([$]?[A-Z]{1,3}[$]?\d+(?::[$]?[A-Z]{1,3}[$]?\d+)?)")
        for m in pattern_local.finditer(formula):
            refs.append((current_sheet, m.group(1)))

    out = []
    seen = set()
    for r in refs:
        if r not in seen:
            seen.add(r)
            out.append(r)
    return out

def strip_refs(formula: str) -> str:
    x = formula
    x = re.sub(r"(?:'[^']+'|[A-Za-z0-9_\. ]+)![$]?[A-Z]{1,3}[$]?\d+(?::[$]?[A-Z]{1,3}[$]?\d+)?", " ", x)
    x = re.sub(r"[$]?[A-Z]{1,3}[$]?\d+(?::[$]?[A-Z]{1,3}[$]?\d+)?", " ", x)
    return x

def extract_constants(formula: str) -> list[str]:
    if not formula or not formula.startswith("="):
        return []
    cleaned = strip_refs(formula)
    vals = re.findall(r"[-+]?\d+(?:\.\d+)?", cleaned)
    out = []
    seen = set()
    for v in vals:
        if v in {"0", "1", "+0", "+1", "-0", "-1"}:
            continue
        if v not in seen:
            seen.add(v)
            out.append(v)
    return out

def write_csv(path: Path, rows: list[dict], columns: list[str]):
    df = pd.DataFrame(rows, columns=columns)
    df.to_csv(path, index=False, encoding="utf-8-sig")

def main():
    ensure_dirs()
    wb_path = get_latest_working_copy()
    wb = load_workbook(filename=wb_path, data_only=False, keep_links=True)
    wb_values = load_workbook(filename=wb_path, data_only=True, keep_links=True)

    inputs_rows = []
    outputs_rows = []
    formulas_rows = []
    const_rows = []
    dep_rows = []

    target_present = [s for s in TARGET_SHEETS if s in wb.sheetnames]

    for sheet_name in target_present:
        ws = wb[sheet_name]
        ws_values = wb_values[sheet_name]

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue

                coord = cell.coordinate
                value = cell.value
                value_calc = ws_values[coord].value
                formula = value if isinstance(value, str) and value.startswith("=") else ""

                left = ws.cell(row=cell.row, column=cell.column - 1).value if cell.column > 1 else None
                up = ws.cell(row=cell.row - 1, column=cell.column).value if cell.row > 1 else None
                context = f"{safe_str(left)} | {safe_str(up)}"

                # INPUTS CORE
                if value is not None and not formula:
                    if text_has_keyword(context, KEYWORDS_INPUT):
                        inputs_rows.append({
                            "hoja": sheet_name,
                            "celda": coord,
                            "valor": safe_str(value),
                            "contexto": context,
                            "tipo": "INPUT_CANDIDATO_CORE"
                        })

                # OUTPUTS CORE
                if value is not None and not formula:
                    if text_has_keyword(left, KEYWORDS_OUTPUT) or text_has_keyword(up, KEYWORDS_OUTPUT):
                        outputs_rows.append({
                            "hoja": sheet_name,
                            "celda": coord,
                            "valor": safe_str(value),
                            "etiqueta_izquierda": safe_str(left),
                            "etiqueta_superior": safe_str(up),
                            "tipo": "OUTPUT_CANDIDATO_CORE"
                        })

                # FORMULAS CORE
                if formula:
                    refs = extract_refs(formula, sheet_name)
                    local_complex = (
                        formula.count("(") + formula.count(",") + formula.count("+") +
                        formula.count("-") + formula.count("*") + formula.count("/") +
                        formula.count(":")
                    )

                    formulas_rows.append({
                        "hoja": sheet_name,
                        "celda": coord,
                        "formula": formula,
                        "valor_calculado": safe_str(value_calc),
                        "referencias": " ; ".join(f"{s}!{a}" for s, a in refs[:30]),
                        "complejidad": local_complex
                    })

                    for c in extract_constants(formula):
                        const_rows.append({
                            "hoja": sheet_name,
                            "celda": coord,
                            "formula": formula,
                            "constante": c,
                            "observacion": "CONST_EMBEBIDA_CORE"
                        })

                    for s, a in refs:
                        dep_rows.append({
                            "origen": f"{s}!{a}",
                            "destino": f"{sheet_name}!{coord}",
                            "tipo_relacion": "CORE_FORMULA_REF"
                        })

    # dedupe
    def dedupe(rows, key_fields):
        out = []
        seen = set()
        for r in rows:
            key = tuple(r[k] for k in key_fields)
            if key not in seen:
                seen.add(key)
                out.append(r)
        return out

    inputs_rows = dedupe(inputs_rows, ["hoja", "celda"])
    outputs_rows = dedupe(outputs_rows, ["hoja", "celda"])
    formulas_rows = dedupe(formulas_rows, ["hoja", "celda"])
    const_rows = dedupe(const_rows, ["hoja", "celda", "constante"])
    dep_rows = dedupe(dep_rows, ["origen", "destino"])

    write_csv(CORE_INPUTS, inputs_rows, ["hoja", "celda", "valor", "contexto", "tipo"])
    write_csv(CORE_OUTPUTS, outputs_rows, ["hoja", "celda", "valor", "etiqueta_izquierda", "etiqueta_superior", "tipo"])
    write_csv(CORE_FORMULAS, formulas_rows, ["hoja", "celda", "formula", "valor_calculado", "referencias", "complejidad"])
    write_csv(CORE_CONSTANTES, const_rows, ["hoja", "celda", "formula", "constante", "observacion"])
    write_csv(CORE_DEPENDENCIAS, dep_rows, ["origen", "destino", "tipo_relacion"])

    summary = {
        "timestamp_utc": now_utc(),
        "workbook": str(wb_path),
        "target_sheets_present": target_present,
        "core_inputs_count": len(inputs_rows),
        "core_outputs_count": len(outputs_rows),
        "core_formulas_count": len(formulas_rows),
        "core_constants_count": len(const_rows),
        "core_dependencies_count": len(dep_rows)
    }

    with CORE_SUMMARY.open("w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print(json.dumps(summary, ensure_ascii=False, indent=2))

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {type(e).__name__}: {e}", file=sys.stderr)
        sys.exit(1)
