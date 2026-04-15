from __future__ import annotations

import json
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


PROJECT_ROOT = Path(r"C:\01. GitHub\BeSmart15dModel")
DATA_DIR = PROJECT_ROOT / "01_DATA"
CLONE_DIR = PROJECT_ROOT / "08_CLONE"
EXPORTS_DIR = PROJECT_ROOT / "09_EXPORTS"
LOGS_DIR = PROJECT_ROOT / "06_TOOLS" / "Logs"

DB_PATH = DATA_DIR / "BeSmart15dModel.stage1.sqlite"
SUMMARY_JSON = EXPORTS_DIR / "CLONE.STAGE3.FLUJO_MENSUAL.SUMMARY.json"
RUN_LOG = LOGS_DIR / "rebuild_flujo_mensual_ia_stage3_from_sqlite.log"

TARGET_SHEET = "Flujo mensual.IA"

# stage 3 assumptions
OCCUPANCY_RATE = 0.90
MONTHLY_EXPENSE_PLACEHOLDER = 250000.0

# block to rebuild
START_ROW = 2
END_ROW = 24
START_COL = 1
END_COL = 6


def now_utc() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def log(msg: str) -> None:
    with RUN_LOG.open("a", encoding="utf-8") as f:
        f.write(f"[{now_utc()}] {msg}\n")


def reset_log() -> None:
    RUN_LOG.write_text("", encoding="utf-8")


def get_latest_stage2_clone() -> Path:
    files = sorted(CLONE_DIR.glob("*.stage2.IA.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"No se encontró ningún archivo *.stage2.IA.xlsx en {CLONE_DIR}")
    return files[0]


def parse_float(value):
    if value is None:
        return None
    txt = str(value).strip()
    if txt == "":
        return None
    try:
        txt = txt.replace(",", "")
        return float(txt)
    except Exception:
        return None


def read_rent_base_from_sqlite() -> dict:
    if not DB_PATH.exists():
        raise FileNotFoundError(f"No existe DB: {DB_PATH}")

    conn = sqlite3.connect(DB_PATH)
    try:
        render_cells = pd.read_sql_query("""
            SELECT sheet_name, cell_address, value, formula, number_format, data_type
            FROM render_cells
            WHERE sheet_name = 'Arriendo'
        """, conn)

        if render_cells.empty:
            raise ValueError("No hay registros de render_cells para hoja Arriendo")

        # solo valores literales, no fórmulas
        df = render_cells.copy()
        df["formula"] = df["formula"].fillna("")
        df = df[df["formula"] == ""]

        # parse numérico
        df["value_num"] = df["value"].apply(parse_float)
        df = df[df["value_num"].notna()]

        # filtro pragmático: positivos y no absurdamente pequeños
        df = df[(df["value_num"] > 0) & (df["value_num"] >= 1000)]

        if df.empty:
            raise ValueError("No se encontraron valores numéricos literales razonables en Arriendo")

        # estrategia stage 3:
        # tomar el máximo valor numérico literal como proxy de ingreso mensual base
        best = df.sort_values(["value_num", "cell_address"], ascending=[False, True]).iloc[0]

        return {
            "monthly_rent_base": float(best["value_num"]),
            "source_sheet": str(best["sheet_name"]),
            "source_cell": str(best["cell_address"]),
            "source_method": "MAX_NUMERIC_LITERAL_IN_ARRIENDO"
        }
    finally:
        conn.close()


def compute_stage3_rows(monthly_rent_base: float):
    monthly_income_full = monthly_rent_base
    monthly_income_effective = monthly_rent_base * OCCUPANCY_RATE
    monthly_expense = MONTHLY_EXPENSE_PLACEHOLDER
    monthly_net = monthly_income_effective - monthly_expense

    rows = [
        ["STAGE3 - SQLITE + ARRIENDO INPUT", "", "", "", "", ""],
        ["Concepto", "Valor", "Unidad", "Origen", "Comentario", "Estado"],
        ["Ingreso mensual full ocupación", monthly_income_full, "CLP", "SQLITE/ARRIENDO", "Base inferida desde hoja Arriendo", "OK"],
        ["Ocupación estabilizada", OCCUPANCY_RATE, "ratio", "PARAMETER", "Supuesto estabilizado confirmado", "OK"],
        ["Ingreso mensual efectivo", monthly_income_effective, "CLP", "ENGINE_STAGE3", "Ingreso full x ocupación", "OK"],
        ["Egreso mensual placeholder", monthly_expense, "CLP", "ENGINE_STAGE3", "Placeholder técnico temporal", "PENDING_REFINEMENT"],
        ["Flujo mensual neto", monthly_net, "CLP", "ENGINE_STAGE3", "Ingreso efectivo - egreso", "OK"],
        ["Flujo acumulado 1 mes", monthly_net, "CLP", "ENGINE_STAGE3", "Acumulado simple", "OK"],
        ["Ingreso anual estabilizado", monthly_income_effective * 12, "CLP", "ENGINE_STAGE3", "Mensual efectivo x 12", "OK"],
        ["Nota", "Este bloque usa base inferida desde SQLite / Arriendo.", "", "SYSTEM", "", "INFO"],
    ]
    return rows


def clear_block(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).value = None


def style_stage_block(ws, start_row: int, row_count: int) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    info_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    for c in range(1, 7):
        ws.cell(row=start_row, column=c).fill = header_fill
        ws.cell(row=start_row + 1, column=c).fill = header_fill

    ws.cell(row=start_row + row_count - 1, column=1).fill = info_fill
    ws.cell(row=start_row + row_count - 1, column=2).fill = info_fill


def write_rows(ws, start_row: int, rows: list[list]) -> None:
    for i, row in enumerate(rows):
        excel_row = start_row + i
        for j, value in enumerate(row, start=1):
            ws.cell(row=excel_row, column=j).value = value


def apply_number_formats(ws, start_row: int) -> None:
    currency_rows = [start_row + 2, start_row + 4, start_row + 5, start_row + 6, start_row + 7, start_row + 8]
    ratio_rows = [start_row + 3]

    for r in currency_rows:
        ws.cell(row=r, column=2).number_format = '#,##0.00'

    for r in ratio_rows:
        ws.cell(row=r, column=2).number_format = '0.00%'


def main() -> int:
    reset_log()
    log("Inicio rebuild_flujo_mensual_ia_stage3_from_sqlite.")

    try:
        source_path = get_latest_stage2_clone()
        log(f"Stage2 clone seleccionado: {source_path}")

        rent_info = read_rent_base_from_sqlite()
        monthly_rent_base = rent_info["monthly_rent_base"]
        log(f"Base inferida desde SQLite: {monthly_rent_base} ({rent_info['source_sheet']}!{rent_info['source_cell']})")

        wb = load_workbook(source_path)

        if TARGET_SHEET not in wb.sheetnames:
            raise ValueError(f"No existe la hoja objetivo: {TARGET_SHEET}")

        ws = wb[TARGET_SHEET]

        rows = compute_stage3_rows(monthly_rent_base)
        clear_block(ws, START_ROW, END_ROW, START_COL, END_COL)
        write_rows(ws, START_ROW, rows)
        style_stage_block(ws, START_ROW, len(rows))
        apply_number_formats(ws, START_ROW)

        out_name = source_path.name.replace(".stage2.IA.xlsx", ".stage3.IA.xlsx")
        out_path = CLONE_DIR / out_name
        wb.save(out_path)

        summary = {
            "timestamp_utc": now_utc(),
            "source_stage2_clone": str(source_path),
            "output_stage3_clone": str(out_path),
            "target_sheet": TARGET_SHEET,
            "rebuild_block": {
                "start_row": START_ROW,
                "end_row": END_ROW,
                "start_col": START_COL,
                "end_col": END_COL
            },
            "input_inference": rent_info,
            "assumptions": {
                "occupancy_rate": OCCUPANCY_RATE,
                "monthly_expense_placeholder": MONTHLY_EXPENSE_PLACEHOLDER
            },
            "status": "OK_STAGE3_FLUJO_MENSUAL_REBUILT_FROM_SQLITE"
        }

        with SUMMARY_JSON.open("w", encoding="utf-8") as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)

        log(f"Workbook stage3 guardado: {out_path}")
        log("Proceso completado OK.")
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0

    except Exception as e:
        log(f"ERROR: {type(e).__name__}: {e}")
        print(f"ERROR: {type(e).__name__}: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
