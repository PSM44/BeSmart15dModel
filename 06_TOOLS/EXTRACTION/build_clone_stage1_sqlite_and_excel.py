from __future__ import annotations

import json
import shutil
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\01. GitHub\BeSmart15dModel")
WORKING_COPY_DIR = PROJECT_ROOT / "01_RAW" / "WORKING_COPY"
CORE_DIR = PROJECT_ROOT / "02_AUDIT" / "CORE"
RENDER_DIR = PROJECT_ROOT / "02_AUDIT" / "RENDER_METADATA"
CLONE_DIR = PROJECT_ROOT / "08_CLONE"
LOGS_DIR = PROJECT_ROOT / "06_TOOLS" / "Logs"
EXPORTS_DIR = PROJECT_ROOT / "09_EXPORTS"
DATA_DIR = PROJECT_ROOT / "01_DATA"

DB_PATH = DATA_DIR / "BeSmart15dModel.stage1.sqlite"
SUMMARY_JSON = EXPORTS_DIR / "CLONE.STAGE1.SUMMARY.json"
RUN_LOG = LOGS_DIR / "build_clone_stage1_sqlite_and_excel.log"

TARGET_SHEETS = [
    "Resumen",
    "IVA Dptos",
    "Hoja1",
    "Arriendo",
    "Flujo mensual",
    "Flujo 5 años C12",
    "Flujo 5 años C12 (2)",
    "Crédito Comercial 12A",
    "Crédito C12",
    "Flujo 5 años C20",
    "Crédito Comercial 20A",
    "Crédito C20",
]

CSV_TABLE_MAP = {
    "CORE.INPUTS.csv": "core_inputs",
    "CORE.OUTPUTS.csv": "core_outputs",
    "CORE.CONSTANTES.csv": "core_constantes",
    "CORE.DEPENDENCIAS.csv": "core_dependencias",
    "RENDER.SHEETS.csv": "render_sheets",
    "RENDER.CELLS.csv": "render_cells",
    "RENDER.MERGES.csv": "render_merges",
    "RENDER.COLUMNS.csv": "render_columns",
    "RENDER.ROWS.csv": "render_rows",
    "RENDER.FREEZE_PANES.csv": "render_freeze_panes",
    "RENDER.PRINT_SETUP.csv": "render_print_setup",
}


def now_utc() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def ensure_dirs() -> None:
    for p in [CLONE_DIR, LOGS_DIR, EXPORTS_DIR, DATA_DIR]:
        p.mkdir(parents=True, exist_ok=True)


def log(msg: str) -> None:
    with RUN_LOG.open("a", encoding="utf-8") as f:
        f.write(f"[{now_utc()}] {msg}\n")


def get_latest_working_copy() -> Path:
    files = sorted(WORKING_COPY_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"No se encontraron .xlsx en {WORKING_COPY_DIR}")
    return files[0]


def import_csv_if_exists(conn: sqlite3.Connection, csv_path: Path, table_name: str, imported: list[str], missing: list[str]) -> None:
    if csv_path.exists():
        df = pd.read_csv(csv_path, encoding="utf-8-sig")
        df.to_sql(table_name, conn, if_exists="replace", index=False)
        imported.append(f"{table_name} <- {csv_path.name} ({len(df)} rows)")
        log(f"CSV importado: {csv_path} -> {table_name} ({len(df)} rows)")
    else:
        missing.append(str(csv_path))
        log(f"CSV faltante: {csv_path}")


def reset_log() -> None:
    RUN_LOG.write_text("", encoding="utf-8")


def make_stage1_clone(source_wb_path: Path) -> dict:
    wb = load_workbook(filename=source_wb_path, data_only=False, keep_links=True)

    existing_names = set(wb.sheetnames)
    created_ia_sheets = []

    for sheet_name in list(wb.sheetnames):
        if sheet_name not in TARGET_SHEETS:
            continue

        ia_name = f"{sheet_name}.IA"
        if ia_name in existing_names:
            log(f"Hoja ya existía y no se duplicó: {ia_name}")
            continue

        src_ws = wb[sheet_name]
        cloned_ws = wb.copy_worksheet(src_ws)
        cloned_ws.title = ia_name
        created_ia_sheets.append(ia_name)
        existing_names.add(ia_name)
        log(f"Hoja duplicada: {sheet_name} -> {ia_name}")

    out_name = f"{source_wb_path.stem}.stage1.IA.xlsx"
    out_path = CLONE_DIR / out_name
    wb.save(out_path)
    log(f"Workbook clon stage1 guardado: {out_path}")

    return {
        "clone_workbook_path": str(out_path),
        "created_ia_sheets": created_ia_sheets,
        "created_ia_sheet_count": len(created_ia_sheets),
    }


def create_db() -> dict:
    if DB_PATH.exists():
        DB_PATH.unlink()

    conn = sqlite3.connect(DB_PATH)
    imported = []
    missing = []

    try:
        for filename, table_name in CSV_TABLE_MAP.items():
            if filename.startswith("CORE."):
                csv_path = CORE_DIR / filename
            else:
                csv_path = RENDER_DIR / filename
            import_csv_if_exists(conn, csv_path, table_name, imported, missing)

        # metadata mínima adicional
        meta = pd.DataFrame([
            {"meta_key": "project_root", "meta_value": str(PROJECT_ROOT)},
            {"meta_key": "db_created_utc", "meta_value": now_utc()},
        ])
        meta.to_sql("project_metadata", conn, if_exists="replace", index=False)

        conn.commit()
    finally:
        conn.close()

    return {
        "db_path": str(DB_PATH),
        "tables_imported": imported,
        "missing_csv": missing,
    }


def main() -> int:
    ensure_dirs()
    reset_log()
    log("Inicio build_clone_stage1_sqlite_and_excel.")

    try:
        source_wb = get_latest_working_copy()
        log(f"Working copy seleccionada: {source_wb}")

        db_info = create_db()
        clone_info = make_stage1_clone(source_wb)

        summary = {
            "timestamp_utc": now_utc(),
            "source_workbook": str(source_wb),
            "db_path": db_info["db_path"],
            "tables_imported_count": len(db_info["tables_imported"]),
            "tables_imported": db_info["tables_imported"],
            "missing_csv_count": len(db_info["missing_csv"]),
            "missing_csv": db_info["missing_csv"],
            "clone_workbook_path": clone_info["clone_workbook_path"],
            "created_ia_sheet_count": clone_info["created_ia_sheet_count"],
            "created_ia_sheets": clone_info["created_ia_sheets"],
            "stage": "STAGE1_SQLITE_AND_TEMPLATE_CLONE"
        }

        with SUMMARY_JSON.open("w", encoding="utf-8") as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)

        log("Proceso completado OK.")
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0

    except Exception as e:
        log(f"ERROR: {type(e).__name__}: {e}")
        print(f"ERROR: {type(e).__name__}: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
