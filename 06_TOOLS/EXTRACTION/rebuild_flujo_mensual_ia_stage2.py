from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


PROJECT_ROOT = Path(r"C:\01. GitHub\BeSmart15dModel")
CLONE_DIR = PROJECT_ROOT / "08_CLONE"
EXPORTS_DIR = PROJECT_ROOT / "09_EXPORTS"
LOGS_DIR = PROJECT_ROOT / "06_TOOLS" / "Logs"

SUMMARY_JSON = EXPORTS_DIR / "CLONE.STAGE2.FLUJO_MENSUAL.SUMMARY.json"
RUN_LOG = LOGS_DIR / "rebuild_flujo_mensual_ia_stage2.log"

TARGET_SHEET = "Flujo mensual.IA"

# ============================================================
# STAGE 2 CONTROLLED ASSUMPTIONS
# ============================================================
MONTHLY_RENT_BASE = 1000000.0
OCCUPANCY_RATE = 0.90
MONTHLY_EXPENSE_PLACEHOLDER = 250000.0

# block to rebuild
START_ROW = 2
END_ROW = 24
START_COL = 1   # A
END_COL = 6     # F


def now_utc() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def log(msg: str) -> None:
    with RUN_LOG.open("a", encoding="utf-8") as f:
        f.write(f"[{now_utc()}] {msg}\n")


def reset_log() -> None:
    RUN_LOG.write_text("", encoding="utf-8")


def get_latest_stage1_clone() -> Path:
    files = sorted(CLONE_DIR.glob("*.stage1.IA.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"No se encontró ningún archivo *.stage1.IA.xlsx en {CLONE_DIR}")
    return files[0]


def compute_stage2_rows():
    monthly_income_full = MONTHLY_RENT_BASE
    monthly_income_effective = MONTHLY_RENT_BASE * OCCUPANCY_RATE
    monthly_expense = MONTHLY_EXPENSE_PLACEHOLDER
    monthly_net = monthly_income_effective - monthly_expense

    rows = [
        ["STAGE2 - REBUILD CONTROLLED BLOCK", "", "", "", "", ""],
        ["Concepto", "Valor", "Unidad", "Origen", "Comentario", "Estado"],
        ["Ingreso mensual full ocupación", monthly_income_full, "CLP", "ENGINE_STAGE2", "Base controlada temporal", "OK"],
        ["Ocupación estabilizada", OCCUPANCY_RATE, "ratio", "ENGINE_STAGE2", "Supuesto estabilizado", "OK"],
        ["Ingreso mensual efectivo", monthly_income_effective, "CLP", "ENGINE_STAGE2", "Ingreso full x ocupación", "OK"],
        ["Egreso mensual placeholder", monthly_expense, "CLP", "ENGINE_STAGE2", "Placeholder técnico", "PENDING_REFINEMENT"],
        ["Flujo mensual neto", monthly_net, "CLP", "ENGINE_STAGE2", "Ingreso efectivo - egreso", "OK"],
        ["Flujo acumulado 1 mes", monthly_net, "CLP", "ENGINE_STAGE2", "Acumulado simple", "OK"],
        ["Ingreso anual estabilizado", monthly_income_effective * 12, "CLP", "ENGINE_STAGE2", "Mensual efectivo x 12", "OK"],
        ["Nota", "Este bloque reemplaza temporalmente parte de la hoja .IA para probar desacople del Excel original.", "", "SYSTEM", "", "INFO"],
    ]
    return rows


def clear_block(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).value = None


def style_stage2_block(ws, start_row: int, row_count: int) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    info_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    for c in range(1, 7):
        ws.cell(row=start_row, column=c).fill = header_fill
        ws.cell(row=start_row + 1, column=c).fill = header_fill

    ws.cell(row=start_row + row_count - 1, column=1).fill = info_fill
    ws.cell(row=start_row + row_count - 1, column=2).fill = info_fill


def write_rows(ws, start_row: int, rows: list[list]) -> None:
    for i, row in enumerate(rows):
        excel_row = start_row + i
        for j, value in enumerate(row, start=1):
            ws.cell(row=excel_row, column=j).value = value


def apply_number_formats(ws, start_row: int, row_count: int) -> None:
    currency_rows = [start_row + 2, start_row + 4, start_row + 5, start_row + 6, start_row + 7, start_row + 8]
    ratio_rows = [start_row + 3]

    for r in currency_rows:
        ws.cell(row=r, column=2).number_format = '#,##0.00'

    for r in ratio_rows:
        ws.cell(row=r, column=2).number_format = '0.00%'


def main() -> int:
    reset_log()
    log("Inicio rebuild_flujo_mensual_ia_stage2.")

    try:
        source_path = get_latest_stage1_clone()
        log(f"Stage1 clone seleccionado: {source_path}")

        wb = load_workbook(source_path)

        if TARGET_SHEET not in wb.sheetnames:
            raise ValueError(f"No existe la hoja objetivo: {TARGET_SHEET}")

        ws = wb[TARGET_SHEET]

        rows = compute_stage2_rows()
        clear_block(ws, START_ROW, END_ROW, START_COL, END_COL)
        write_rows(ws, START_ROW, rows)
        style_stage2_block(ws, START_ROW, len(rows))
        apply_number_formats(ws, START_ROW, len(rows))

        out_name = source_path.name.replace(".stage1.IA.xlsx", ".stage2.IA.xlsx")
        out_path = CLONE_DIR / out_name
        wb.save(out_path)

        summary = {
            "timestamp_utc": now_utc(),
            "source_stage1_clone": str(source_path),
            "output_stage2_clone": str(out_path),
            "target_sheet": TARGET_SHEET,
            "rebuild_block": {
                "start_row": START_ROW,
                "end_row": END_ROW,
                "start_col": START_COL,
                "end_col": END_COL
            },
            "assumptions": {
                "monthly_rent_base": MONTHLY_RENT_BASE,
                "occupancy_rate": OCCUPANCY_RATE,
                "monthly_expense_placeholder": MONTHLY_EXPENSE_PLACEHOLDER
            },
            "status": "OK_STAGE2_FLUJO_MENSUAL_REBUILT"
        }

        with SUMMARY_JSON.open("w", encoding="utf-8") as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)

        log(f"Workbook stage2 guardado: {out_path}")
        log("Proceso completado OK.")
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0

    except Exception as e:
        log(f"ERROR: {type(e).__name__}: {e}")
        print(f"ERROR: {type(e).__name__}: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
