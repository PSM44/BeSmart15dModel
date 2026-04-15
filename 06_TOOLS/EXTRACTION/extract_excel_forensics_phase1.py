from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(r"C:\01. GitHub\BeSmart15dModel")
WORKING_COPY_DIR = PROJECT_ROOT / "01_RAW" / "WORKING_COPY"
TABLES_DIR = PROJECT_ROOT / "02_AUDIT" / "TABLES"
LOGS_DIR = PROJECT_ROOT / "02_AUDIT" / "LOGS"
ANALYSIS_DIR = PROJECT_ROOT / "03_ANALYSIS"
EXPORTS_DIR = PROJECT_ROOT / "09_EXPORTS"

AUDIT_LOG_PATH = LOGS_DIR / "AUDIT.LOG.txt"
RUN_LOG_PATH = LOGS_DIR / "EXTRACT.PHASE1.LOG.txt"
SUMMARY_JSON_PATH = EXPORTS_DIR / "EXTRACT.PHASE1.SUMMARY.json"

VOLATILE_FUNCTIONS = {
    "NOW", "TODAY", "RAND", "RANDBETWEEN", "OFFSET", "INDIRECT", "INFO", "CELL"
}

OUTPUT_KEYWORDS = [
    "tir", "irr", "van", "npv", "vpn", "rentabilidad", "retorno", "utilidad",
    "flujo", "payback", "roi", "roe", "cap rate", "resultado", "resumen",
    "ingresos", "egresos", "valor final", "valor venta", "caja"
]

INPUT_KEYWORDS = [
    "supuesto", "input", "entrada", "parámetro", "parametro", "tasa", "vacancia",
    "precio", "canon", "arriendo", "cap rate", "descuento", "costo", "gasto",
    "comisión", "comision", "impuesto", "pie", "bono", "crédito", "credito"
]

ERROR_TOKENS = {
    "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!"
}


@dataclass
class SheetInventoryRow:
    hoja: str
    estado: str
    descripcion: str
    criticidad: str


@dataclass
class CriticalCellRow:
    hoja: str
    celda: str
    valor_mostrado: str
    valor_real: str
    formula: str
    clasificacion: str
    criticidad: str
    dependencia_input: str
    dependencia_output: str


@dataclass
class DependencyRow:
    origen: str
    destino: str
    tipo_relacion: str


@dataclass
class InputRow:
    hoja: str
    celda: str
    tipo: str
    validacion: str
    observacion: str


@dataclass
class OutputRow:
    hoja: str
    celda: str
    descripcion: str
    criticidad: str


@dataclass
class ConstantRow:
    hoja: str
    celda: str
    formula: str
    constante_detectada: str
    impacto: str


@dataclass
class VolatileRow:
    hoja: str
    celda: str
    funcion: str
    impacto: str


@dataclass
class ErrorRow:
    hoja: str
    celda: str
    tipo_error: str
    impacto: str


@dataclass
class FlowRow:
    hoja_origen: str
    hoja_destino: str
    tipo_flujo: str
    descripcion: str


@dataclass
class RiskRow:
    tipo: str
    ubicacion: str
    descripcion: str
    impacto: str


@dataclass
class RankingSheetRow:
    hoja: str
    criticidad: str
    justificacion: str


@dataclass
class RankingOutputRow:
    output: str
    ubicacion: str
    criticidad: str
    justificacion: str


@dataclass
class RankingComponentRow:
    componente: str
    ubicacion: str
    riesgo: str
    justificacion: str


def now_utc_str() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def ensure_dirs() -> None:
    for p in [TABLES_DIR, LOGS_DIR, ANALYSIS_DIR, EXPORTS_DIR]:
        p.mkdir(parents=True, exist_ok=True)


def append_log(path: Path, message: str) -> None:
    with path.open("a", encoding="utf-8") as f:
        f.write(f"[{now_utc_str()}] {message}\n")


def safe_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def normalize_text(value: Any) -> str:
    return safe_str(value).strip().lower()


def is_error_value(value: Any) -> bool:
    txt = safe_str(value).strip().upper()
    return txt in ERROR_TOKENS


def classify_manual_value(value: Any) -> str:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "INPUT_MANUAL_NUMERICO"
    return "INPUT_MANUAL_ALFANUMERICO"


def count_formula_complexity(formula: str) -> int:
    if not formula:
        return 0
    return (
        formula.count("(")
        + formula.count(",")
        + formula.count("+")
        + formula.count("-")
        + formula.count("*")
        + formula.count("/")
        + formula.count(":")
    )


def classify_sheet_state(ws) -> str:
    state = safe_str(getattr(ws, "sheet_state", "visible"))
    return state or "visible"


def used_range_text(ws) -> str:
    col = get_column_letter(ws.max_column) if ws.max_column else "A"
    row = ws.max_row if ws.max_row else 1
    return f"A1:{col}{row}"


def has_data_validation(ws, coord: str) -> bool:
    try:
        dvs = ws.data_validations
        if dvs is None or not getattr(dvs, "dataValidation", None):
            return False
        for dv in dvs.dataValidation:
            for rng in dv.ranges.ranges:
                if coord in rng:
                    return True
        return False
    except Exception:
        return False


def get_validation_desc(ws, coord: str) -> str:
    try:
        dvs = ws.data_validations
        if dvs is None or not getattr(dvs, "dataValidation", None):
            return ""
        found = []
        for dv in dvs.dataValidation:
            for rng in dv.ranges.ranges:
                if coord in rng:
                    found.append(
                        f"type={safe_str(dv.type)};operator={safe_str(dv.operator)};formula1={safe_str(dv.formula1)};formula2={safe_str(dv.formula2)}"
                    )
        return " | ".join(found)
    except Exception:
        return ""


def text_has_keyword(text: str, keywords: Iterable[str]) -> bool:
    t = normalize_text(text)
    return any(k in t for k in keywords)


def is_potential_output_cell(ws, cell) -> bool:
    if cell.value is None:
        return False
    txt = normalize_text(cell.value)
    if text_has_keyword(txt, OUTPUT_KEYWORDS):
        if cell.column < ws.max_column:
            right = ws.cell(row=cell.row, column=cell.column + 1)
            if right.value is not None:
                return True
    return False


def extract_formula_references(formula: str, current_sheet: str) -> list[tuple[str, str]]:
    refs: list[tuple[str, str]] = []
    if not formula or not formula.startswith("="):
        return refs

    pattern_with_sheet = re.compile(
        r"(?:'([^']+)'|([A-Za-z0-9_\.]+))!([$]?[A-Z]{1,3}[$]?\d+(?::[$]?[A-Z]{1,3}[$]?\d+)?)"
    )
    for m in pattern_with_sheet.finditer(formula):
        sheet = m.group(1) or m.group(2) or current_sheet
        ref = m.group(3)
        refs.append((sheet, ref))

    pattern_local = re.compile(r"(?<![A-Za-z0-9_])([$]?[A-Z]{1,3}[$]?\d+(?::[$]?[A-Z]{1,3}[$]?\d+)?)")
    if "!" not in formula:
        for m in pattern_local.finditer(formula):
            refs.append((current_sheet, m.group(1)))

    deduped = []
    seen = set()
    for item in refs:
        if item not in seen:
            seen.add(item)
            deduped.append(item)
    return deduped


def strip_cell_references_for_constant_scan(formula: str) -> str:
    if not formula:
        return ""
    cleaned = formula

    cleaned = re.sub(
        r"(?:'[^']+'|[A-Za-z0-9_\.]+)![$]?[A-Z]{1,3}[$]?\d+(?::[$]?[A-Z]{1,3}[$]?\d+)?",
        " ",
        cleaned,
    )
    cleaned = re.sub(
        r"[$]?[A-Z]{1,3}[$]?\d+(?::[$]?[A-Z]{1,3}[$]?\d+)?",
        " ",
        cleaned,
    )
    return cleaned


def extract_embedded_constants(formula: str) -> list[str]:
    if not formula or not formula.startswith("="):
        return []

    cleaned = strip_cell_references_for_constant_scan(formula)
    tokens = re.findall(r"[-+]?\d+(?:\.\d+)?", cleaned)

    result = []
    seen = set()
    for token in tokens:
        if token in {"0", "1", "+0", "+1", "-0", "-1"}:
            continue
        if token not in seen:
            seen.add(token)
            result.append(token)
    return result


def extract_volatile_functions(formula: str) -> list[str]:
    if not formula or not formula.startswith("="):
        return []
    upper = formula.upper()
    found = []
    for fn in VOLATILE_FUNCTIONS:
        if f"{fn}(" in upper:
            found.append(fn)
    return found


def estimate_criticidad_for_sheet(ws) -> str:
    score = 0
    name = normalize_text(ws.title)

    if text_has_keyword(name, OUTPUT_KEYWORDS):
        score += 2
    if ws.max_row > 100 or ws.max_column > 20:
        score += 1

    formula_count = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_count += 1
                if formula_count > 50:
                    score += 2
                    break
        if formula_count > 50:
            break

    if score >= 4:
        return "ALTA"
    if score >= 2:
        return "MEDIA"
    return "BAJA"


def estimate_risk_formula(formula: str) -> str:
    if not formula:
        return "BAJO"

    complexity = count_formula_complexity(formula)
    vol = extract_volatile_functions(formula)
    consts = extract_embedded_constants(formula)
    external = "[" in formula and "]" in formula

    score = 0
    if complexity >= 8:
        score += 2
    elif complexity >= 4:
        score += 1
    if vol:
        score += 2
    if consts:
        score += 1
    if external:
        score += 2

    if score >= 4:
        return "ALTO"
    if score >= 2:
        return "MEDIO"
    return "BAJO"


def get_latest_working_copy() -> Path:
    files = sorted(WORKING_COPY_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"No se encontraron .xlsx en {WORKING_COPY_DIR}")
    return files[0]


def dataframe_to_csv(path: Path, rows: list[dict], columns: list[str]) -> None:
    df = pd.DataFrame(rows, columns=columns)
    df.to_csv(path, index=False, encoding="utf-8-sig")


def get_defined_names_safe(wb) -> list[dict[str, str]]:
    result = []
    try:
        items = list(wb.defined_names.items())
        for name, defn in items:
            attr_text = safe_str(getattr(defn, "attr_text", ""))
            local_sheet_id = safe_str(getattr(defn, "localSheetId", ""))
            hidden = safe_str(getattr(defn, "hidden", ""))
            result.append({
                "name": safe_str(name),
                "attr_text": attr_text,
                "localSheetId": local_sheet_id,
                "hidden": hidden,
            })
        return result
    except Exception:
        pass

    try:
        values = list(wb.defined_names.values())
        for defn in values:
            result.append({
                "name": safe_str(getattr(defn, "name", "")),
                "attr_text": safe_str(getattr(defn, "attr_text", "")),
                "localSheetId": safe_str(getattr(defn, "localSheetId", "")),
                "hidden": safe_str(getattr(defn, "hidden", "")),
            })
        return result
    except Exception:
        return []


def main() -> int:
    ensure_dirs()
    append_log(RUN_LOG_PATH, "Inicio extractor Phase 1.")

    try:
        workbook_path = get_latest_working_copy()
        append_log(RUN_LOG_PATH, f"Working copy seleccionada: {workbook_path}")

        wb = load_workbook(filename=workbook_path, data_only=False, keep_links=True)
        wb_values = load_workbook(filename=workbook_path, data_only=True, keep_links=True)

        inventory_sheets: list[SheetInventoryRow] = []
        critical_cells: list[CriticalCellRow] = []
        dependencies: list[DependencyRow] = []
        inputs_rows: list[InputRow] = []
        outputs_rows: list[OutputRow] = []
        constants_rows: list[ConstantRow] = []
        volatile_rows: list[VolatileRow] = []
        error_rows: list[ErrorRow] = []
        flow_rows: list[FlowRow] = []
        risk_rows: list[RiskRow] = []
        ranking_sheet_rows: list[RankingSheetRow] = []
        ranking_output_rows: list[RankingOutputRow] = []
        ranking_component_rows: list[RankingComponentRow] = []

        sheet_out_refs: set[str] = set()
        dependency_pairs: set[tuple[str, str, str]] = set()
        flow_pairs: set[tuple[str, str, str, str]] = set()

        formula_count = 0
        validation_count = 0
        merged_ranges_count = 0
        external_links_count = 0
        comments_count = 0

        external_links_obj = getattr(wb, "_external_links", None)
        if external_links_obj:
            external_links_count = len(external_links_obj)

        defined_names_serialized = get_defined_names_safe(wb)
        defined_names_count = len(defined_names_serialized)

        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    if is_potential_output_cell(ws, cell):
                        right = ws.cell(row=cell.row, column=cell.column + 1) if cell.column < ws.max_column else None
                        if right and right.value is not None:
                            sheet_out_refs.add(f"{ws.title}!{right.coordinate}")
                            outputs_rows.append(OutputRow(
                                hoja=ws.title,
                                celda=right.coordinate,
                                descripcion=f"Etiqueta izquierda: {safe_str(cell.value)}",
                                criticidad="ALTA"
                            ))

        for ws in wb.worksheets:
            ws_values = wb_values[ws.title]
            state = classify_sheet_state(ws)
            criticidad_sheet = estimate_criticidad_for_sheet(ws)

            desc = (
                f"state={state}; used_range={used_range_text(ws)}; "
                f"rows={ws.max_row}; cols={ws.max_column}; "
                f"tables={len(getattr(ws, 'tables', {}))}; "
                f"merged_ranges={len(ws.merged_cells.ranges)}"
            )
            inventory_sheets.append(SheetInventoryRow(
                hoja=ws.title,
                estado=state,
                descripcion=desc,
                criticidad=criticidad_sheet
            ))
            ranking_sheet_rows.append(RankingSheetRow(
                hoja=ws.title,
                criticidad=criticidad_sheet,
                justificacion=desc
            ))

            merged_ranges_count += len(ws.merged_cells.ranges)

            if ws.data_validations and getattr(ws.data_validations, "dataValidation", None):
                validation_count += len(ws.data_validations.dataValidation)

            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.comment:
                        comments_count += 1

            sheet_to_sheet_refs: defaultdict[str, int] = defaultdict(int)

            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue

                    coord = cell.coordinate
                    value = cell.value
                    value_only = ws_values[coord].value

                    formula = value if isinstance(value, str) and value.startswith("=") else ""
                    display = safe_str(value_only if value_only is not None else value)
                    real = safe_str(value)

                    if formula:
                        formula_count += 1

                        refs = extract_formula_references(formula, ws.title)
                        for ref_sheet, ref_addr in refs:
                            dependency_pairs.add((
                                f"{ref_sheet}!{ref_addr}",
                                f"{ws.title}!{coord}",
                                "formula_reference"
                            ))
                            sheet_to_sheet_refs[ref_sheet] += 1

                        embedded_constants = extract_embedded_constants(formula)
                        for const in embedded_constants:
                            constants_rows.append(ConstantRow(
                                hoja=ws.title,
                                celda=coord,
                                formula=formula,
                                constante_detectada=const,
                                impacto=estimate_risk_formula(formula)
                            ))

                        vol_fns = extract_volatile_functions(formula)
                        for fn in vol_fns:
                            volatile_rows.append(VolatileRow(
                                hoja=ws.title,
                                celda=coord,
                                funcion=fn,
                                impacto=estimate_risk_formula(formula)
                            ))

                        risk_level = estimate_risk_formula(formula)
                        if risk_level != "BAJO":
                            ranking_component_rows.append(RankingComponentRow(
                                componente="FORMULA",
                                ubicacion=f"{ws.title}!{coord}",
                                riesgo=risk_level,
                                justificacion=formula
                            ))
                            risk_rows.append(RiskRow(
                                tipo="FORMULA_COMPLEJA",
                                ubicacion=f"{ws.title}!{coord}",
                                descripcion=formula,
                                impacto=risk_level
                            ))

                    if is_error_value(value_only):
                        error_rows.append(ErrorRow(
                            hoja=ws.title,
                            celda=coord,
                            tipo_error=safe_str(value_only),
                            impacto="ALTO"
                        ))
                        risk_rows.append(RiskRow(
                            tipo="ERROR_EXCEL",
                            ubicacion=f"{ws.title}!{coord}",
                            descripcion=f"Valor error detectado: {safe_str(value_only)}",
                            impacto="ALTO"
                        ))

                    if has_data_validation(ws, coord) and value is not None:
                        inputs_rows.append(InputRow(
                            hoja=ws.title,
                            celda=coord,
                            tipo="CONTROL",
                            validacion=get_validation_desc(ws, coord),
                            observacion="Celda con validación"
                        ))

                    if value is not None and not formula:
                        cell_class = classify_manual_value(value)
                        left = ws.cell(row=cell.row, column=cell.column - 1).value if cell.column > 1 else None
                        up = ws.cell(row=cell.row - 1, column=cell.column).value if cell.row > 1 else None
                        context = f"{safe_str(left)} | {safe_str(up)}"

                        if has_data_validation(ws, coord) or text_has_keyword(context, INPUT_KEYWORDS):
                            inputs_rows.append(InputRow(
                                hoja=ws.title,
                                celda=coord,
                                tipo=cell_class,
                                validacion=get_validation_desc(ws, coord),
                                observacion=f"Contexto input: {context}".strip()
                            ))

                    is_critical = False
                    clasificacion = "NO_CONFIRMADO"
                    criticidad = "BAJA"
                    dep_input = ""
                    dep_output = ""

                    if f"{ws.title}!{coord}" in sheet_out_refs:
                        is_critical = True
                        clasificacion = "OUTPUT"
                        criticidad = "ALTA"

                    elif formula:
                        refs = extract_formula_references(formula, ws.title)
                        if refs or count_formula_complexity(formula) >= 4:
                            is_critical = True
                            clasificacion = "CALCULO"
                            criticidad = "ALTA" if estimate_risk_formula(formula) == "ALTO" else "MEDIA"
                            dep_input = "; ".join(f"{s}!{r}" for s, r in refs[:20])

                    elif value is not None and has_data_validation(ws, coord):
                        is_critical = True
                        clasificacion = "CONTROL"
                        criticidad = "MEDIA"

                    elif value is not None:
                        left = ws.cell(row=cell.row, column=cell.column - 1).value if cell.column > 1 else None
                        up = ws.cell(row=cell.row - 1, column=cell.column).value if cell.row > 1 else None
                        context = f"{safe_str(left)} | {safe_str(up)}"
                        if text_has_keyword(context, INPUT_KEYWORDS):
                            is_critical = True
                            clasificacion = classify_manual_value(value)
                            criticidad = "MEDIA"

                    if is_critical:
                        if f"{ws.title}!{coord}" in sheet_out_refs:
                            dep_output = f"{ws.title}!{coord}"
                        critical_cells.append(CriticalCellRow(
                            hoja=ws.title,
                            celda=coord,
                            valor_mostrado=display,
                            valor_real=real,
                            formula=formula,
                            clasificacion=clasificacion,
                            criticidad=criticidad,
                            dependencia_input=dep_input,
                            dependencia_output=dep_output
                        ))

            for ref_sheet, count in sheet_to_sheet_refs.items():
                if ref_sheet != ws.title:
                    flow_type = "REFERENCIA"
                    if count >= 20:
                        flow_type = "AGREGACION"
                    elif count >= 5:
                        flow_type = "TRANSFORMACION"
                    flow_pairs.add((
                        ref_sheet,
                        ws.title,
                        flow_type,
                        f"{count} referencias detectadas desde fórmulas"
                    ))

        unique_outputs = {}
        for row in outputs_rows:
            key = (row.hoja, row.celda)
            if key not in unique_outputs:
                unique_outputs[key] = row
        outputs_rows = list(unique_outputs.values())

        for src, dst, rel in sorted(dependency_pairs):
            dependencies.append(DependencyRow(
                origen=src,
                destino=dst,
                tipo_relacion=rel
            ))

        for src_sheet, dst_sheet, tf, desc in sorted(flow_pairs):
            flow_rows.append(FlowRow(
                hoja_origen=src_sheet,
                hoja_destino=dst_sheet,
                tipo_flujo=tf,
                descripcion=desc
            ))

        risk_rows.append(RiskRow(
            tipo="LIMITACION_HERRAMIENTA",
            ubicacion="WORKBOOK",
            descripcion="No se auditó VBA / eventos / Power Query / Power Pivot / ODBC-OLEDB / objetos complejos en esta fase Python base.",
            impacto="ALTO"
        ))

        if external_links_count > 0:
            risk_rows.append(RiskRow(
                tipo="ENLACE_EXTERNO",
                ubicacion="WORKBOOK",
                descripcion=f"Se detectaron {external_links_count} external links.",
                impacto="ALTO"
            ))

        if defined_names_count == 0:
            risk_rows.append(RiskRow(
                tipo="NO_CONFIRMADO",
                ubicacion="WORKBOOK",
                descripcion="No se detectaron nombres definidos o la carga no los expuso con claridad.",
                impacto="MEDIO"
            ))

        for out in outputs_rows:
            ranking_output_rows.append(RankingOutputRow(
                output=f"{out.hoja}!{out.celda}",
                ubicacion=f"{out.hoja}!{out.celda}",
                criticidad=out.criticidad,
                justificacion=out.descripcion
            ))

        for row in volatile_rows:
            ranking_component_rows.append(RankingComponentRow(
                componente="FUNCION_VOLATIL",
                ubicacion=f"{row.hoja}!{row.celda}",
                riesgo=row.impacto,
                justificacion=row.funcion
            ))

        for row in error_rows:
            ranking_component_rows.append(RankingComponentRow(
                componente="ERROR_EXCEL",
                ubicacion=f"{row.hoja}!{row.celda}",
                riesgo=row.impacto,
                justificacion=row.tipo_error
            ))

        dataframe_to_csv(
            TABLES_DIR / "11.01_INVENTARIO_DE_HOJAS.csv",
            [asdict(x) for x in inventory_sheets],
            ["hoja", "estado", "descripcion", "criticidad"]
        )
        dataframe_to_csv(
            TABLES_DIR / "11.02_INVENTARIO_DE_CELDAS_CRITICAS.csv",
            [asdict(x) for x in critical_cells],
            ["hoja", "celda", "valor_mostrado", "valor_real", "formula", "clasificacion", "criticidad", "dependencia_input", "dependencia_output"]
        )
        dataframe_to_csv(
            TABLES_DIR / "11.03_MATRIZ_DE_DEPENDENCIAS.csv",
            [asdict(x) for x in dependencies],
            ["origen", "destino", "tipo_relacion"]
        )
        dataframe_to_csv(
            TABLES_DIR / "11.04_INVENTARIO_DE_INPUTS.csv",
            [asdict(x) for x in inputs_rows],
            ["hoja", "celda", "tipo", "validacion", "observacion"]
        )
        dataframe_to_csv(
            TABLES_DIR / "11.05_INVENTARIO_DE_OUTPUTS_CLAVE.csv",
            [asdict(x) for x in outputs_rows],
            ["hoja", "celda", "descripcion", "criticidad"]
        )
        dataframe_to_csv(
            TABLES_DIR / "11.06_INVENTARIO_DE_CONSTANTES_EMBEBIDAS.csv",
            [asdict(x) for x in constants_rows],
            ["hoja", "celda", "formula", "constante_detectada", "impacto"]
        )
        dataframe_to_csv(
            TABLES_DIR / "11.07_INVENTARIO_DE_FUNCIONES_VOLATILES.csv",
            [asdict(x) for x in volatile_rows],
            ["hoja", "celda", "funcion", "impacto"]
        )
        dataframe_to_csv(
            TABLES_DIR / "11.08_INVENTARIO_DE_ERRORES.csv",
            [asdict(x) for x in error_rows],
            ["hoja", "celda", "tipo_error", "impacto"]
        )
        dataframe_to_csv(
            TABLES_DIR / "11.09_MAPA_DE_HOJAS_Y_FLUJO_GENERAL.csv",
            [asdict(x) for x in flow_rows],
            ["hoja_origen", "hoja_destino", "tipo_flujo", "descripcion"]
        )
        dataframe_to_csv(
            TABLES_DIR / "11.10_LISTA_DE_RIESGOS_TECNICOS_Y_FINANCIEROS.csv",
            [asdict(x) for x in risk_rows],
            ["tipo", "ubicacion", "descripcion", "impacto"]
        )
        dataframe_to_csv(
            TABLES_DIR / "11.11_RANKING_DE_CRITICIDAD_POR_HOJA.csv",
            [asdict(x) for x in ranking_sheet_rows],
            ["hoja", "criticidad", "justificacion"]
        )
        dataframe_to_csv(
            TABLES_DIR / "11.12_RANKING_DE_CRITICIDAD_POR_OUTPUT.csv",
            [asdict(x) for x in ranking_output_rows],
            ["output", "ubicacion", "criticidad", "justificacion"]
        )
        dataframe_to_csv(
            TABLES_DIR / "11.13_RANKING_DE_RIESGO_POR_COMPONENTE.csv",
            [asdict(x) for x in ranking_component_rows],
            ["componente", "ubicacion", "riesgo", "justificacion"]
        )

        summary = {
            "timestamp_utc": now_utc_str(),
            "project_root": str(PROJECT_ROOT),
            "workbook_name": workbook_path.name,
            "workbook_path": str(workbook_path),
            "sheet_count": len(wb.worksheets),
            "formula_count": formula_count,
            "error_count": len(error_rows),
            "defined_names_count": defined_names_count,
            "validation_count": validation_count,
            "merged_ranges_count": merged_ranges_count,
            "external_links_count": external_links_count,
            "comments_count": comments_count,
            "outputs_detected_count": len(outputs_rows),
            "inputs_detected_count": len(inputs_rows),
            "critical_cells_count": len(critical_cells),
            "dependencies_count": len(dependencies),
            "limitations": [
                "No cobertura completa de VBA/eventos.",
                "No cobertura completa de Power Query/Power Pivot.",
                "No cobertura completa de objetos visuales complejos.",
                "Heurísticas de OUTPUT e INPUT requieren validación humana."
            ],
            "defined_names": defined_names_serialized,
        }

        with SUMMARY_JSON_PATH.open("w", encoding="utf-8") as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)

        append_log(AUDIT_LOG_PATH, "Extractor Phase 1 ejecutado.")
        append_log(AUDIT_LOG_PATH, f"Workbook analizado: {summary['workbook_name']}")
        append_log(AUDIT_LOG_PATH, f"Hojas detectadas: {summary['sheet_count']}")
        append_log(AUDIT_LOG_PATH, f"Formulas detectadas: {summary['formula_count']}")
        append_log(AUDIT_LOG_PATH, f"Errores detectados: {summary['error_count']}")
        append_log(AUDIT_LOG_PATH, f"Nombres definidos detectados: {summary['defined_names_count']}")
        append_log(AUDIT_LOG_PATH, f"Validaciones detectadas: {summary['validation_count']}")
        append_log(AUDIT_LOG_PATH, f"Merges detectados: {summary['merged_ranges_count']}")
        append_log(AUDIT_LOG_PATH, f"Links externos detectados: {summary['external_links_count']}")

        append_log(RUN_LOG_PATH, "Extractor Phase 1 completado OK.")
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        sys.exit(0)

    except Exception as e:
        ensure_dirs()
        append_log(RUN_LOG_PATH, f"ERROR: {type(e).__name__}: {e}")
        append_log(AUDIT_LOG_PATH, f"ERROR extractor Phase 1: {type(e).__name__}: {e}")
        print(f"ERROR: {type(e).__name__}: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
