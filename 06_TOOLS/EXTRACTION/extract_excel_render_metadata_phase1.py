from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

PROJECT_ROOT = Path(r"C:\01. GitHub\BeSmart15dModel")
WORKING_COPY_DIR = PROJECT_ROOT / "01_RAW" / "WORKING_COPY"
LOGS_DIR = PROJECT_ROOT / "06_TOOLS" / "Logs"
META_DIR = PROJECT_ROOT / "02_AUDIT" / "RENDER_METADATA"
EXPORTS_DIR = PROJECT_ROOT / "09_EXPORTS"

SHEETS_CSV = META_DIR / "RENDER.SHEETS.csv"
CELLS_CSV = META_DIR / "RENDER.CELLS.csv"
MERGES_CSV = META_DIR / "RENDER.MERGES.csv"
COLS_CSV = META_DIR / "RENDER.COLUMNS.csv"
ROWS_CSV = META_DIR / "RENDER.ROWS.csv"
FREEZE_CSV = META_DIR / "RENDER.FREEZE_PANES.csv"
PRINT_CSV = META_DIR / "RENDER.PRINT_SETUP.csv"
SUMMARY_JSON = EXPORTS_DIR / "RENDER.EXTRACT.SUMMARY.json"

TARGET_SHEETS = [
    "Resumen",
    "IVA Dptos",
    "Hoja1",
    "Arriendo",
    "Flujo mensual",
    "Flujo 5 años C12",
    "Flujo 5 años C12 (2)",
    "Crédito Comercial 12A",
    "Crédito C12",
    "Flujo 5 años C20",
    "Crédito Comercial 20A",
    "Crédito C20",
]

def now_utc():
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

def safe_str(v):
    if v is None:
        return ""
    return str(v)

def ensure_dirs():
    for p in [LOGS_DIR, META_DIR, EXPORTS_DIR]:
        p.mkdir(parents=True, exist_ok=True)

def get_latest_working_copy():
    files = sorted(WORKING_COPY_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"No se encontraron .xlsx en {WORKING_COPY_DIR}")
    return files[0]

def style_key(cell):
    try:
        return safe_str(cell.style_id)
    except Exception:
        return ""

def main():
    ensure_dirs()
    wb_path = get_latest_working_copy()
    wb = load_workbook(filename=wb_path, data_only=False, keep_links=True)

    sheets_rows = []
    cells_rows = []
    merges_rows = []
    cols_rows = []
    rows_rows = []
    freeze_rows = []
    print_rows = []

    target_present = [s for s in TARGET_SHEETS if s in wb.sheetnames]

    for order_idx, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        if sheet_name not in TARGET_SHEETS:
            continue

        used_range = f"A1:{ws.max_column and ws.max_row and ws.calculate_dimension().split(':')[-1] or 'A1'}"
        sheets_rows.append({
            "sheet_name_original": sheet_name,
            "sheet_name_ia": f"{sheet_name}.IA",
            "sheet_order": order_idx,
            "sheet_state": safe_str(getattr(ws, "sheet_state", "visible")),
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "used_range": safe_str(ws.calculate_dimension()),
            "tab_color": safe_str(getattr(getattr(ws.sheet_properties, 'tabColor', None), 'rgb', "")),
        })

        for merge_range in ws.merged_cells.ranges:
            merges_rows.append({
                "sheet_name": sheet_name,
                "merge_range": safe_str(merge_range)
            })

        for col_letter, dim in ws.column_dimensions.items():
            cols_rows.append({
                "sheet_name": sheet_name,
                "col_letter": col_letter,
                "width": safe_str(dim.width),
                "hidden": safe_str(dim.hidden),
                "bestFit": safe_str(getattr(dim, "bestFit", "")),
            })

        for row_idx, dim in ws.row_dimensions.items():
            rows_rows.append({
                "sheet_name": sheet_name,
                "row_num": row_idx,
                "height": safe_str(dim.height),
                "hidden": safe_str(dim.hidden),
            })

        freeze_rows.append({
            "sheet_name": sheet_name,
            "freeze_panes": safe_str(ws.freeze_panes)
        })

        print_rows.append({
            "sheet_name": sheet_name,
            "print_area": safe_str(ws.print_area),
            "print_title_rows": safe_str(ws.print_title_rows),
            "print_title_cols": safe_str(ws.print_title_cols),
            "orientation": safe_str(getattr(ws.page_setup, "orientation", "")),
            "paperSize": safe_str(getattr(ws.page_setup, "paperSize", "")),
            "fitToWidth": safe_str(getattr(ws.page_setup, "fitToWidth", "")),
            "fitToHeight": safe_str(getattr(ws.page_setup, "fitToHeight", "")),
        })

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue

                cells_rows.append({
                    "sheet_name": sheet_name,
                    "cell_address": cell.coordinate,
                    "row_num": cell.row,
                    "col_num": cell.column,
                    "value": safe_str(cell.value),
                    "formula": safe_str(cell.value) if isinstance(cell.value, str) and cell.value.startswith("=") else "",
                    "data_type": safe_str(cell.data_type),
                    "number_format": safe_str(cell.number_format),
                    "style_id": style_key(cell),
                    "font_name": safe_str(getattr(cell.font, "name", "")),
                    "font_size": safe_str(getattr(cell.font, "sz", "")),
                    "font_bold": safe_str(getattr(cell.font, "b", "")),
                    "font_italic": safe_str(getattr(cell.font, "i", "")),
                    "fill_type": safe_str(getattr(cell.fill, "fill_type", "")),
                    "fill_fgColor": safe_str(getattr(getattr(cell.fill, "fgColor", None), "rgb", "")),
                    "align_h": safe_str(getattr(cell.alignment, "horizontal", "")),
                    "align_v": safe_str(getattr(cell.alignment, "vertical", "")),
                    "wrap_text": safe_str(getattr(cell.alignment, "wrap_text", "")),
                    "border_left": safe_str(getattr(getattr(cell.border, "left", None), "style", "")),
                    "border_right": safe_str(getattr(getattr(cell.border, "right", None), "style", "")),
                    "border_top": safe_str(getattr(getattr(cell.border, "top", None), "style", "")),
                    "border_bottom": safe_str(getattr(getattr(cell.border, "bottom", None), "style", "")),
                    "is_merged_anchor": "TRUE" if any(cell.coordinate == str(rng).split(":")[0] for rng in ws.merged_cells.ranges) else "FALSE",
                })

    pd.DataFrame(sheets_rows).to_csv(SHEETS_CSV, index=False, encoding="utf-8-sig")
    pd.DataFrame(cells_rows).to_csv(CELLS_CSV, index=False, encoding="utf-8-sig")
    pd.DataFrame(merges_rows).to_csv(MERGES_CSV, index=False, encoding="utf-8-sig")
    pd.DataFrame(cols_rows).to_csv(COLS_CSV, index=False, encoding="utf-8-sig")
    pd.DataFrame(rows_rows).to_csv(ROWS_CSV, index=False, encoding="utf-8-sig")
    pd.DataFrame(freeze_rows).to_csv(FREEZE_CSV, index=False, encoding="utf-8-sig")
    pd.DataFrame(print_rows).to_csv(PRINT_CSV, index=False, encoding="utf-8-sig")

    summary = {
        "timestamp_utc": now_utc(),
        "workbook": str(wb_path),
        "target_sheets_present": target_present,
        "sheet_count": len(sheets_rows),
        "cell_rows_count": len(cells_rows),
        "merge_rows_count": len(merges_rows),
        "column_rows_count": len(cols_rows),
        "row_rows_count": len(rows_rows),
        "freeze_rows_count": len(freeze_rows),
        "print_rows_count": len(print_rows),
    }

    with SUMMARY_JSON.open("w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print(json.dumps(summary, ensure_ascii=False, indent=2))

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {type(e).__name__}: {e}", file=sys.stderr)
        sys.exit(1)
